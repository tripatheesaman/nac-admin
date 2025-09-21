
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView
from django.views import View
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from psycopg2.extras import RealDictCursor
from decouple import config
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import logging
import os
import pandas as pd
import openpyxl
import psycopg2

logger = logging.getLogger(__name__)

from .models import ProcessedFile, StaffDetails
from .forms import FileUploadForm, ProcessingOptionsForm, StaffDetailsForm, StaffFilterForm
from .services import ExcelProcessorService


def get_department_filtered_queryset(user, model_class):
    """
    Helper function to get department-filtered querysets.
    Superusers see all data, regular users see only their department's data.
    """
    if user.is_superuser:
        return model_class.objects.all()
    elif user.department:
        if model_class == ProcessedFile:
            # For files, filter by users in the same department
            return model_class.objects.filter(user__department=user.department)
        elif model_class == StaffDetails:
            # For staff, filter by department
            return model_class.objects.filter(department=user.department)
    else:
        # User has no department, return empty queryset
        return model_class.objects.none()


# Progress polling endpoint for AJAX
@login_required(login_url='/app/login/')
@require_GET
def get_progress(request, progress_id):
    """Return dummy progress info for file processing (expand as needed)"""
    try:
        processed_file = ProcessedFile.objects.get(id=progress_id)
        if processed_file.status == 'completed':
            return JsonResponse({
                'progress': 100,
                'status': 'completed',
                'message': 'Processing complete.'
            })
        elif processed_file.status == 'failed':
            return JsonResponse({
                'progress': 100,
                'status': 'failed',
                'message': processed_file.error_message or 'Processing failed.'
            })
        elif processed_file.status == 'processing':
            return JsonResponse({
                'progress': 50,
                'status': 'processing',
                'message': 'Processing in progress...'
            })
        else:
            return JsonResponse({
                'progress': 10,
                'status': 'pending',
                'message': 'Waiting to start processing...'
            })
    except ProcessedFile.DoesNotExist:
        return JsonResponse({
            'progress': 100,
            'status': 'failed',
            'message': 'File not found.'
        }, status=404)
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, DetailView
from django.views import View
from django.http import JsonResponse, HttpResponse, Http404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
import os
import pandas as pd
import openpyxl
import psycopg2
from django.conf import settings
from psycopg2.extras import RealDictCursor
from decouple import config
import logging
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

logger = logging.getLogger(__name__)

from .models import ProcessedFile
from .forms import FileUploadForm, ProcessingOptionsForm, StaffDetailsForm, StaffFilterForm
from .services import ExcelProcessorService


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class HomeView(TemplateView):
    """Home view for file upload and processing"""
    template_name = 'processor/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['upload_form'] = FileUploadForm()
        context['options_form'] = ProcessingOptionsForm()
        
        # Get department-filtered queryset
        files_queryset = get_department_filtered_queryset(self.request.user, ProcessedFile)
        context['recent_files'] = files_queryset.select_related('user').order_by('-created_at')[:5]
        
        # Add file statistics (department-filtered)
        context['total_files'] = files_queryset.count()
        context['completed_files'] = files_queryset.filter(status='completed').count()
        context['pending_files'] = files_queryset.filter(status='pending').count()
        context['processing_files'] = files_queryset.filter(status='processing').count()
        context['failed_files'] = files_queryset.filter(status='failed').count()
        
        return context
    
    def post(self, request, *args, **kwargs):
        from django.http import JsonResponse
        upload_form = FileUploadForm(request.POST, request.FILES)
        options_form = ProcessingOptionsForm(request.POST, request.FILES)

        if upload_form.is_valid() and options_form.is_valid():
            processed_file = upload_form.save(commit=False)
            processed_file.user = request.user
            processed_file.status = 'pending'
            processed_file.save()

            # For AJAX: return progress_id for polling
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                # Use the processed_file id as progress_id
                return JsonResponse({
                    'success': True,
                    'progress_id': processed_file.id
                })

            # Non-AJAX fallback: process immediately (legacy)
            try:
                service = ExcelProcessorService()
                input_path = processed_file.original_file.path
                output_filename = f"processed_{os.path.basename(input_path)}"
                from django.conf import settings
                output_path = os.path.join(settings.MEDIA_ROOT, 'processed', output_filename)
                
                logger.info(f"MEDIA_ROOT: {settings.MEDIA_ROOT}")
                logger.info(f"Output path: {output_path}")
                logger.info(f"Directory to create: {os.path.dirname(output_path)}")
                
                # Ensure the processed directory exists
                try:
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)
                    logger.info(f"Ensured directory exists: {os.path.dirname(output_path)}")
                    
                    # Verify the directory was created
                    if os.path.exists(os.path.dirname(output_path)):
                        logger.info(f"Directory verification successful: {os.path.dirname(output_path)}")
                        
                        # Test if we can write to the directory
                        test_file = os.path.join(os.path.dirname(output_path), '.test_write')
                        try:
                            with open(test_file, 'w') as f:
                                f.write('test')
                            os.remove(test_file)
                            logger.info(f"Directory write test successful: {os.path.dirname(output_path)}")
                        except Exception as write_error:
                            logger.error(f"Directory write test failed: {os.path.dirname(output_path)} - {write_error}")
                            raise Exception(f"Directory is not writable: {os.path.dirname(output_path)}")
                    else:
                        logger.error(f"Directory verification failed: {os.path.dirname(output_path)}")
                        raise Exception(f"Directory was not created: {os.path.dirname(output_path)}")
                        
                except Exception as e:
                    logger.error(f"Failed to create directory {os.path.dirname(output_path)}: {e}")
                    raise

                template_path = None
                if options_form.cleaned_data['output_type'] == 'template':
                    template_file = options_form.cleaned_data['template_file']
                    if template_file:
                        template_path = template_file.path

                logger.info(f"Processing file: input_path={input_path}, output_path={output_path}")
                result = service.process_file(input_path, output_path, template_path)

                if result['success']:
                    processed_file.status = 'completed'
                    processed_file.processed_file = os.path.join('processed', output_filename)
                    processed_file.save()
                    messages.success(request, f"File processed successfully! {result['output_rows']} records processed.")
                else:
                    processed_file.status = 'failed'
                    processed_file.error_message = result.get('error', 'Unknown error')
                    processed_file.save()
                    messages.error(request, f"Processing failed: {result.get('error', 'Unknown error')}")

            except Exception as e:
                processed_file.status = 'failed'
                processed_file.error_message = str(e)
                processed_file.save()
                messages.error(request, f"Processing failed: {str(e)}")

            return redirect('processor:home')

        # If invalid, return errors for AJAX or render for normal
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            errors = {}
            errors.update(upload_form.errors)
            errors.update(options_form.errors)
            return JsonResponse({'success': False, 'errors': errors})

        context = self.get_context_data()
        context['upload_form'] = upload_form
        context['options_form'] = options_form
        return render(request, self.template_name, context)


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class FileListView(ListView):
    """List all processed files"""
    model = ProcessedFile
    template_name = 'processor/file_list.html'
    context_object_name = 'files'
    paginate_by = 20
    
    def get_queryset(self):
        files_queryset = get_department_filtered_queryset(self.request.user, ProcessedFile)
        return files_queryset.select_related('user').order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add additional context for file statistics (department-filtered)
        files_queryset = get_department_filtered_queryset(self.request.user, ProcessedFile)
        context['total_files'] = files_queryset.count()
        context['completed_files'] = files_queryset.filter(status='completed').count()
        context['pending_files'] = files_queryset.filter(status='pending').count()
        context['processing_files'] = files_queryset.filter(status='processing').count()
        context['failed_files'] = files_queryset.filter(status='failed').count()
        return context


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class FileDetailView(DetailView):
    """Show details of a specific processed file"""
    model = ProcessedFile
    template_name = 'processor/file_detail.html'
    context_object_name = 'file'
    pk_url_kwarg = 'file_id'
    
    def get_queryset(self):
        return get_department_filtered_queryset(self.request.user, ProcessedFile).select_related('user')


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class DownloadView(View):
    """Download processed file"""
    def get(self, request, file_id):
        files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
        processed_file = get_object_or_404(files_queryset, id=file_id)
        
        if processed_file.processed_file and processed_file.status == 'completed':
            file_path = processed_file.processed_file.path
            if os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                    return response
        
        messages.error(request, "File not found or not processed yet.")
        return redirect('processor:file_detail', file_id=file_id)


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class ProcessFileView(View):
    """AJAX endpoint for processing files"""
    def post(self, request, file_id):
        try:
            files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
            processed_file = get_object_or_404(files_queryset, id=file_id)
            
            if processed_file.status == 'completed':
                return JsonResponse({
                    'success': False,
                    'message': 'File already processed'
                })
            
            processed_file.status = 'processing'
            processed_file.save()
            
            service = ExcelProcessorService()
            input_path = processed_file.original_file.path
            output_filename = f"processed_{os.path.basename(input_path)}"
            # Use MEDIA_ROOT for proper path construction
            from django.conf import settings
            output_path = os.path.join(settings.MEDIA_ROOT, 'processed', output_filename)
            
            logger.info(f"MEDIA_ROOT: {settings.MEDIA_ROOT}")
            logger.info(f"Output path: {output_path}")
            logger.info(f"Directory to create: {os.path.dirname(output_path)}")
            
            # Ensure the processed directory exists
            try:
                os.makedirs(os.path.dirname(output_path), exist_ok=True)
                logger.info(f"Ensured directory exists: {os.path.dirname(output_path)}")
                
                # Verify the directory was created
                if os.path.exists(os.path.dirname(output_path)):
                    logger.info(f"Directory verification successful: {os.path.dirname(output_path)}")
                    
                    # Test if we can write to the directory
                    test_file = os.path.join(os.path.dirname(output_path), '.test_write')
                    try:
                        with open(test_file, 'w') as f:
                            f.write('test')
                        os.remove(test_file)
                        logger.info(f"Directory write test successful: {os.path.dirname(output_path)}")
                    except Exception as write_error:
                        logger.error(f"Directory write test failed: {os.path.dirname(output_path)} - {write_error}")
                        raise Exception(f"Directory is not writable: {os.path.dirname(output_path)}")
                else:
                    logger.error(f"Directory verification failed: {os.path.dirname(output_path)}")
                    raise Exception(f"Directory was not created: {os.path.dirname(output_path)}")
                    
            except Exception as e:
                logger.error(f"Failed to create directory {os.path.dirname(output_path)}: {e}")
                raise
            
            logger.info(f"Processing file: input_path={input_path}, output_path={output_path}")
            result = service.process_file(input_path, output_path)
            
            if result['success']:
                processed_file.status = 'completed'
                # Store relative path for Django FileField
                processed_file.processed_file = os.path.join('processed', output_filename)
                processed_file.save()
                return JsonResponse({
                    'success': True,
                    'message': f"File processed successfully! {result['output_rows']} records processed."
                })
            else:
                processed_file.status = 'failed'
                processed_file.error_message = result.get('error', 'Unknown error')
                processed_file.save()
                return JsonResponse({
                    'success': False,
                    'message': f"Processing failed: {result.get('error', 'Unknown error')}"
                })
                
        except Exception as e:
            # Safely handle cases where processed_file may not be set
            try:
                if 'processed_file' in locals() and processed_file:
                    processed_file.status = 'failed'
                    processed_file.error_message = str(e)
                    processed_file.save()
            except Exception:
                pass
            
            return JsonResponse({
                'success': False,
                'message': f"Processing failed: {str(e)}"
            })


@login_required(login_url='/app/login/')
@require_http_methods(["GET", "POST", "DELETE"]) 
def delete_file(request, file_id):
    """Delete a processed file"""
    files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
    processed_file = get_object_or_404(files_queryset, id=file_id)

    # Allow GET to act as a redirect-safe confirmationless delete fallback
    if request.method in ["POST", "DELETE", "GET"]:
        # Delete associated files
        if processed_file.original_file and os.path.exists(processed_file.original_file.path):
            try:
                os.remove(processed_file.original_file.path)
            except Exception:
                pass
        if processed_file.processed_file and os.path.exists(processed_file.processed_file.path):
            try:
                os.remove(processed_file.processed_file.path)
            except Exception:
                pass
        processed_file.delete()
        messages.success(request, "File deleted successfully.")
        # For AJAX requests, return to list via redirect; non-AJAX GET will also redirect
        return redirect('processor:file_list')
    
    return redirect('processor:file_list')


@login_required(login_url='/app/login/')
def preview_data(request, file_id):
    """Preview extracted data from uploaded file with pagination and search"""
    files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
    processed_file = get_object_or_404(files_queryset, id=file_id)
    
    try:
        service = ExcelProcessorService()
        input_path = processed_file.original_file.path
        
        # Extract data without writing to file (try matrix parser first)
        try:
            data = service.processor.process_matrix_attendance(input_path)
        except Exception:
            if 'Report (1).xls' in input_path or service._is_attendance_file(input_path):
                data = service.processor.process_attendance_file(input_path)
            else:
                data = service.reader.read_excel(input_path)
                data = service.processor.process_data(data)
        
        # Convert to list for template
        if hasattr(data, 'to_dict'):
            data_list = data.to_dict('records')
        else:
            data_list = []
        columns = list(data.columns) if hasattr(data, 'columns') else []

        # Search filter
        search_query = request.GET.get('q', '').strip()
        if search_query:
            data_list = [row for row in data_list if any(search_query.lower() in str(row.get(col, '')).lower() for col in columns)]

        # Pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(data_list, 25)  # 25 records per page
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        is_paginated = paginator.num_pages > 1

        context = {
            'file': processed_file,
            'columns': columns,
            'total_records': len(data_list),
            'page_obj': page_obj,
            'is_paginated': is_paginated,
            'search_query': search_query,
        }
        return render(request, 'processor/preview_data.html', context)
    except Exception as e:
        messages.error(request, f"Error previewing data: {str(e)}")
        return redirect('processor:file_detail', file_id=file_id)


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class StaffListView(ListView):
    """List all staff with filtering capabilities"""
    template_name = 'processor/staff_list.html'
    context_object_name = 'staff_list'
    paginate_by = 20
    
    def get_queryset(self):
        from django.db import connection
        
        # Build the base query
        query = "SELECT * FROM staff_details"
        params = []
        conditions = []
        
        # Apply department filtering
        if not self.request.user.is_superuser and self.request.user.department:
            conditions.append("department_id = %s")
            params.append(self.request.user.department.id)
        elif not self.request.user.is_superuser:
            # User has no department, return empty result
            return []
        
        # Apply filters
        filter_form = StaffFilterForm(self.request.GET)
        if filter_form.is_valid():
            name = filter_form.cleaned_data.get('name')
            staffid = filter_form.cleaned_data.get('staffid')
            section = filter_form.cleaned_data.get('section')
            designation = filter_form.cleaned_data.get('designation')
            level = filter_form.cleaned_data.get('level')
            weekly_off = filter_form.cleaned_data.get('weekly_off')
            type_of_employment = filter_form.cleaned_data.get('type_of_employment')
            priority = filter_form.cleaned_data.get('priority')
            
            if name:
                conditions.append("name ILIKE %s")
                params.append(f'%{name}%')
            if staffid:
                conditions.append("staffid ILIKE %s")
                params.append(f'%{staffid}%')
            if section:
                conditions.append("section ILIKE %s")
                params.append(f'%{section}%')
            if designation:
                conditions.append("designation ILIKE %s")
                params.append(f'%{designation}%')
            if level is not None:
                conditions.append("level = %s")
                params.append(level)
            if weekly_off:
                conditions.append("weekly_off = %s")
                params.append(weekly_off)
            if type_of_employment:
                conditions.append("type_of_employment = %s")
                params.append(type_of_employment)
            if priority is not None:
                conditions.append("priority = %s")
                params.append(priority)
        
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        
        query += " ORDER BY staffid"
        
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = StaffFilterForm(self.request.GET)
        return context


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class StaffCreateView(View):
    """Create new staff member"""
    template_name = 'processor/staff_form.html'
    
    def get(self, request):
        form = StaffDetailsForm()
        return render(request, self.template_name, {'form': form, 'action': 'Add'})
    
    def post(self, request):
        from django.db import connection
        
        form = StaffDetailsForm(request.POST)
        if form.is_valid():
            try:
                # Get form data
                staffid = form.cleaned_data['staffid']
                name = form.cleaned_data['name']
                section = form.cleaned_data['section'] or None
                designation = form.cleaned_data['designation'] or None
                weekly_off = form.cleaned_data['weekly_off']
                level = form.cleaned_data['level'] or None
                type_of_employment = form.cleaned_data['type_of_employment']
                priority = form.cleaned_data['priority'] or 1
                
                with connection.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO staff_details (staffid, name, section, designation, weekly_off, level, type_of_employment, priority)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, [staffid, name, section, designation, weekly_off, level, type_of_employment, priority])
                
                messages.success(request, 'Staff member added successfully!')
                return redirect('processor:staff_list')
            except Exception as e:
                messages.error(request, f"Error adding staff member: {str(e)}")
                print(f"Error: {str(e)}")  # Debug print
        else:
            # Debug: print form errors
            print(f"Form errors: {form.errors}")
            messages.error(request, f"Form validation failed: {form.errors}")
        
        return render(request, self.template_name, {'form': form, 'action': 'Add'})


@method_decorator(login_required(login_url='/app/login/'), name='dispatch')
class StaffUpdateView(View):
    """Update existing staff member"""
    template_name = 'processor/staff_form.html'
    
    def get(self, request, staff_id):
        from django.db import connection
        
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM staff_details WHERE id = %s", [staff_id])
            row = cursor.fetchone()
            if not row:
                raise Http404("Staff member not found")
            
            columns = [col[0] for col in cursor.description]
            staff_data = dict(zip(columns, row))
        
        # Create form with initial data and staff_id for validation
        form = StaffDetailsForm(staff_id=staff_id, initial={
            'staffid': staff_data.get('staffid'),
            'name': staff_data.get('name'),
            'section': staff_data.get('section'),
            'designation': staff_data.get('designation'),
            'weekly_off': staff_data.get('weekly_off'),
            'level': staff_data.get('level'),
            'type_of_employment': staff_data.get('type_of_employment'),
            'priority': staff_data.get('priority')
        })
        return render(request, self.template_name, {
            'form': form, 
            'action': 'Edit',
            'staff': staff_data
        })
    
    def post(self, request, staff_id):
        from django.db import connection
        
        form = StaffDetailsForm(request.POST, staff_id=staff_id)
        if form.is_valid():
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        UPDATE staff_details 
                        SET staffid = %s, name = %s, section = %s, designation = %s, 
                            weekly_off = %s, level = %s, type_of_employment = %s, priority = %s, 
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, [
                        form.cleaned_data['staffid'],
                        form.cleaned_data['name'],
                        form.cleaned_data['section'],
                        form.cleaned_data['designation'],
                        form.cleaned_data['weekly_off'],
                        form.cleaned_data['level'],
                        form.cleaned_data['type_of_employment'],
                        form.cleaned_data.get('priority') or 1,
                        staff_id
                    ])
                
                messages.success(request, 'Staff member updated successfully!')
                return redirect('processor:staff_list')
            except Exception as e:
                messages.error(request, f"Error updating staff member: {str(e)}")
        
        # Get staff data for form
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM staff_details WHERE id = %s", [staff_id])
            row = cursor.fetchone()
            if row:
                columns = [col[0] for col in cursor.description]
                staff_data = dict(zip(columns, row))
            else:
                staff_data = {}
        
        return render(request, self.template_name, {
            'form': form, 
            'action': 'Edit',
            'staff': staff_data
        })


@login_required(login_url='/app/login/')
@require_POST
def delete_staff(request, staff_id):
    """Delete staff member"""
    from django.db import connection
    
    try:
        with connection.cursor() as cursor:
            # Get staff name before deleting
            cursor.execute("SELECT name FROM staff_details WHERE id = %s", [staff_id])
            row = cursor.fetchone()
            if not row:
                messages.error(request, "Staff member not found")
                return redirect('processor:staff_list')
            
            staff_name = row[0]
            
            # Delete the staff member
            cursor.execute("DELETE FROM staff_details WHERE id = %s", [staff_id])
            
            messages.success(request, f'Staff member "{staff_name}" deleted successfully!')
    except Exception as e:
        messages.error(request, f"Error deleting staff member: {str(e)}")
    
    return redirect('processor:staff_list')


@login_required(login_url='/app/login/')
def get_detailed_leave_details(request, file_id):
    """Get detailed leave details for all employees with pagination and search"""
    try:
        files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
        processed_file = get_object_or_404(files_queryset, id=file_id)
        import pandas as pd
        if processed_file.processed_file:
            try:
                df = pd.read_excel(processed_file.processed_file.path)
            except FileNotFoundError:
                service = ExcelProcessorService()
                df = service.processor.process_attendance_file(processed_file.original_file.path)
        else:
            service = ExcelProcessorService()
            df = service.processor.process_attendance_file(processed_file.original_file.path)

        # Check if required columns exist
        required_columns = ['Employee_ID', 'Employee_Name', 'Designation', 'Status', 'Date']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return render(request, 'processor/leave_details.html', {
                'error': f"Missing required columns: {missing_columns}",
                'leave_list': [],
                'is_paginated': False,
                'search_query': '',
                'page_obj': None,
                'file': processed_file,
            })

        # Build leave details list
        leave_list = []
        
        # Get department-filtered employee IDs
        if not request.user.is_superuser and request.user.department:
            # Get staff IDs from the user's department
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute("SELECT staffid FROM staff_details WHERE department_id = %s", [request.user.department.id])
                department_staff_ids = [row[0] for row in cursor.fetchall()]
            
            # Filter employee IDs to only include those from the user's department
            available_employee_ids = [emp_id for emp_id in df['Employee_ID'].dropna().unique() if str(emp_id) in department_staff_ids]
        else:
            available_employee_ids = df['Employee_ID'].dropna().unique()
        
        for employee_id in available_employee_ids:
            employee_data = df[df['Employee_ID'] == employee_id]
            if employee_data.empty:
                continue
            try:
                employee_name = employee_data['Employee_Name'].iloc[0] if len(employee_data) > 0 else "Unknown"
                designation = employee_data['Designation'].iloc[0] if len(employee_data) > 0 else "Unknown"
            except (IndexError, KeyError):
                employee_name = "Unknown"
                designation = "Unknown"
            present_days = absent_days = weekly_off_days = allowance_days = sick_leave_days = casual_leave_days = personal_leave_days = substitute_leave_days = duty_leave_days = other_leave_days = 0
            for _, row in employee_data.iterrows():
                status = str(row.get('Status', '')).strip().upper()
                if 'P' in status or 'A *' in status:
                    present_days += 1
                    allowance_days += 1
                elif status == 'A':
                    absent_days += 1
                elif 'WO' in status or 'HO' in status:
                    present_days += 1
                    weekly_off_days += 1
                elif 'SL' in status:
                    sick_leave_days += 1
                elif 'CL' in status:
                    casual_leave_days += 1
                elif 'PL' in status:
                    personal_leave_days += 1
                elif 'SUBSTITUTE' in status or 'SUBL' in status:
                    substitute_leave_days += 1
                elif 'DUTY' in status:
                    duty_leave_days += 1
                    # Also count as other leave for the other_leave_days count
                    other_leave_days += 1
                elif 'L' in status:
                    # Count as other leave for the other_leave_days count
                    other_leave_days += 1
                else:
                    # Count as other leave if status doesn't match any specific leave type (PL, SL, CL, SUBSTITUTE)
                    other_leave_days += 1
            leave_list.append({
                'employee_id': employee_id,
                'employee_name': employee_name,
                'designation': designation,
                'present_days': present_days,
                'absent_days': absent_days,
                'weekly_off_days': weekly_off_days,
                'allowance_days': allowance_days,
                'sick_leave_days': sick_leave_days,
                'casual_leave_days': casual_leave_days,
                'personal_leave_days': personal_leave_days,
                'substitute_leave_days': substitute_leave_days,
                'duty_leave_days': duty_leave_days,
                'other_leave_days': other_leave_days,
            })

        # Search filter
        search_query = request.GET.get('q', '').strip()
        if search_query:
            leave_list = [row for row in leave_list if search_query.lower() in str(row['employee_id']).lower() or search_query.lower() in str(row['employee_name']).lower() or search_query.lower() in str(row['designation']).lower()]

        # Pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(leave_list, 25)
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        is_paginated = paginator.num_pages > 1

        return render(request, 'processor/leave_details.html', {
            'leave_list': page_obj.object_list,
            'is_paginated': is_paginated,
            'search_query': search_query,
            'page_obj': page_obj,
            'file': processed_file,
            'error': None,
        })
    except Exception as e:
        return render(request, 'processor/leave_details.html', {
            'error': str(e),
            'leave_list': [],
            'is_paginated': False,
            'search_query': '',
            'page_obj': None,
            'file': None,
        })


@login_required(login_url='/app/login/')
def generate_segregation_report(request, file_id):
    """Generate staff segregation report by sections"""
    from django.db import connection
    import zipfile
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import pandas as pd
    
    try:
        processed_file = get_object_or_404(ProcessedFile, id=file_id)
        
        # Extract data from the uploaded file
        service = ExcelProcessorService()
        input_path = processed_file.original_file.path
        
        # Get attendance data (matrix parser first)
        try:
            attendance_data = service.processor.process_matrix_attendance(input_path)
        except Exception:
            if 'Report (1).xls' in input_path or service._is_attendance_file(input_path):
                attendance_data = service.processor.process_attendance_file(input_path)
            else:
                attendance_data = service.reader.read_excel(input_path)
                attendance_data = service.processor.process_data(attendance_data)
        
        if attendance_data.empty:
            messages.error(request, "No attendance data found in the file.")
            return redirect('processor:file_detail', file_id=file_id)
        
        # Get unique employee IDs and filter out NaN values
        unique_employee_ids = attendance_data['Employee_ID'].dropna().unique()
        
        # Fetch section, employment type, and priority information from database
        employee_details = {}
        with connection.cursor() as cursor:
            for emp_id in unique_employee_ids:
                # Skip if emp_id is NaN or empty
                if pd.isna(emp_id) or str(emp_id).strip() == '':
                    continue
                    
                cursor.execute("SELECT section, type_of_employment, priority FROM staff_details WHERE staffid = %s AND type_of_employment IN ('permanent', 'contract')", [str(emp_id)])
                row = cursor.fetchone()
                if row:
                    employee_details[emp_id] = {
                        'section': row[0] if row[0] else "Unknown Section",
                        'type_of_employment': row[1] if row[1] else 'monthly wages',
                        'priority': row[2] if row[2] else 999
                    }
                else:
                    employee_details[emp_id] = {
                        'section': "Unknown Section",
                        'type_of_employment': 'monthly wages',
                        'priority': 999
                    }
        
        # Group data by sections
        section_data = {}
        for emp_id in unique_employee_ids:
            # Skip if emp_id is NaN or empty
            if pd.isna(emp_id) or str(emp_id).strip() == '':
                continue
                
            section = employee_details.get(emp_id, {}).get('section', "Unknown Section")
            if section not in section_data:
                section_data[section] = []
            
            # Get all records for this employee
            emp_records = attendance_data[attendance_data['Employee_ID'] == emp_id]
            section_data[section].extend(emp_records.to_dict('records'))
        
        # Create ZIP file with separate workbooks for each section
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for section, records in section_data.items():
                if not records:
                    continue
                
                # Create workbook for this section
                wb = Workbook()
                ws = wb.active
                ws.title = f"{section}_Attendance"
                
                # Get period information from the original file
                period_info = "Unknown"
                try:
                    # Try to extract period from the original file
                    service = ExcelProcessorService()
                    input_path = processed_file.original_file.path
                    if input_path.lower().endswith('.xls'):
                        df = pd.read_excel(input_path, engine='xlrd', header=None)
                    else:
                        df = pd.read_excel(input_path, engine='openpyxl', header=None)
                    
                    if len(df) > 8:
                        period_cell = df.iloc[8, 0]
                        if pd.notna(period_cell):
                            import re
                            raw_text = str(period_cell)
                            period_info = re.sub(r"(?i)^\s*period\s*:?,?\s*", "", raw_text).strip() or "Unknown"
                except:
                    pass
                
                # Write main title (single 'period' wording)
                title_cell = ws.cell(row=1, column=1, value=f"Attendance Record of {section} for the period {period_info}")
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                ws.merge_cells('A1:Z1')  # Merge cells for title
                
                # Get unique dates for column headers
                unique_dates = []
                for record in records:
                    date = record.get('Date', '')
                    if date and date not in unique_dates:
                        unique_dates.append(date)
                unique_dates.sort()
                
                # Write employee info headers (row 3)
                ws.cell(row=3, column=1, value="Emp ID").font = Font(bold=True)
                ws.cell(row=3, column=2, value="Emp Name").font = Font(bold=True)
                ws.cell(row=3, column=3, value="Designation").font = Font(bold=True)
                
                # Write Time header (row 3)
                ws.cell(row=3, column=4, value="Time").font = Font(bold=True)
                
                # Write day headers starting from column 5
                current_col = 5
                import re
                for date in unique_dates:
                    day_name = ""
                    day_number = ""
                    # Prefer Day_Name from a record matching this date
                    for record in records:
                        if record.get('Date') == date:
                            day_name = record.get('Day_Name', '')
                            break
                    # Parse "DD Weekday" from the date string (set once, no duplication)
                    if isinstance(date, str):
                        m = re.match(r"^\s*(\d{1,2})\s+([A-Za-z]+)\s*$", date)
                        if m:
                            day_number = m.group(1)
                            # if day_name not set from record, take from header
                            if not day_name:
                                day_name = m.group(2)
                        else:
                            # Fallback: use the whole string as number, no duplicate name
                            day_number = date
                    header_text = f"{day_number} {day_name}".strip()
                    ws.cell(row=3, column=current_col, value=header_text).font = Font(bold=True)
                    current_col += 1
                

                
                # Group records by employee
                employee_data = {}
                for record in records:
                    emp_id = record.get('Employee_ID', '')
                    if pd.isna(emp_id) or str(emp_id).strip() == '':
                        continue
                    
                    if emp_id not in employee_data:
                        emp_details = employee_details.get(emp_id, {})
                        employee_data[emp_id] = {
                            'name': record.get('Employee_Name', ''),
                            'designation': record.get('Designation', ''),
                            'type_of_employment': emp_details.get('type_of_employment', 'monthly wages'),
                            'priority': emp_details.get('priority', 999),
                            'daily_data': {}
                        }
                    
                    date = record.get('Date', '')
                    if date:
                        employee_data[emp_id]['daily_data'][date] = {
                            'in_time': record.get('InTime', ''),
                            'out_time': record.get('OutTime', ''),
                            'status': record.get('Status', ''),
                            'worked_hours': record.get('WorkedHours', '')
                        }
                
                # Sort employees according to priority first, then employment type
                sorted_employees = sorted(employee_data.items(), key=lambda x: (
                    # First by priority (ascending)
                    x[1].get('priority', 999),
                    # Then by employment type (permanent first, then contract)
                    0 if x[1].get('type_of_employment') == 'permanent' else 
                    1 if x[1].get('type_of_employment') == 'contract' else 2,
                    # Then by staffid numeric part (ascending) - handle text staffids
                    int(''.join(filter(str.isdigit, str(x[0])))) if ''.join(filter(str.isdigit, str(x[0]))) and ''.join(filter(str.isdigit, str(x[0]))).isdigit() else 999999
                ))
                
                # Write employee data - each employee takes 4 rows
                current_row = 5
                for emp_id, emp_info in sorted_employees:
                    # Row 1: In Time
                    ws.cell(row=current_row, column=1, value=emp_id)
                    ws.cell(row=current_row, column=2, value=emp_info['name'])
                    ws.cell(row=current_row, column=3, value=emp_info['designation'])
                    ws.cell(row=current_row, column=4, value="In Time").font = Font(bold=True)
                    
                    # Fill In Time data for each day
                    current_col = 5
                    for date in unique_dates:
                        if date in emp_info['daily_data']:
                            ws.cell(row=current_row, column=current_col, value=emp_info['daily_data'][date]['in_time'])
                        current_col += 1
                    current_row += 1
                    
                    # Row 2: Out Time
                    ws.cell(row=current_row, column=4, value="Out Time").font = Font(bold=True)
                    
                    # Fill Out Time data for each day
                    current_col = 5
                    for date in unique_dates:
                        if date in emp_info['daily_data']:
                            ws.cell(row=current_row, column=current_col, value=emp_info['daily_data'][date]['out_time'])
                        current_col += 1
                    current_row += 1
                    
                    # Row 3: Status
                    ws.cell(row=current_row, column=4, value="Status").font = Font(bold=True)
                    
                    # Fill Status data for each day
                    current_col = 5
                    for date in unique_dates:
                        if date in emp_info['daily_data']:
                            ws.cell(row=current_row, column=current_col, value=emp_info['daily_data'][date]['status'])
                        current_col += 1
                    current_row += 1
                    
                    # Row 4: Worked Hours
                    ws.cell(row=current_row, column=4, value="Worked Hours").font = Font(bold=True)
                    
                    # Fill Worked Hours data for each day
                    current_col = 5
                    for date in unique_dates:
                        if date in emp_info['daily_data']:
                            ws.cell(row=current_row, column=current_col, value=emp_info['daily_data'][date]['worked_hours'])
                        current_col += 1
                    current_row += 1
                
                # Apply formatting and merging
                from openpyxl.styles import Border, Side
                
                # Define border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Header row formatting (row 3)
                for col in range(1, current_col):
                    cell = ws.cell(row=3, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", wrap_text=False)
                    cell.border = thin_border
                
                # Set fixed height for header row
                ws.row_dimensions[3].height = 20
                
                # Merge employee info cells and apply borders
                current_row = 5
                for emp_id, emp_info in sorted_employees:
                    # Merge employee info cells (columns A-C) for 4 rows
                    ws.merge_cells(f'A{current_row}:A{current_row + 3}')  # Emp ID
                    ws.merge_cells(f'B{current_row}:B{current_row + 3}')  # Emp Name
                    ws.merge_cells(f'C{current_row}:C{current_row + 3}')  # Designation
                    
                    # Apply borders and formatting to merged cells
                    for col in range(1, 4):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    
                    # Set fixed row height to prevent word wrap
                    fixed_row_height = 20  # Single line height
                    
                    # Set row heights for all 4 rows of this employee
                    ws.row_dimensions[current_row].height = fixed_row_height
                    ws.row_dimensions[current_row + 1].height = fixed_row_height
                    ws.row_dimensions[current_row + 2].height = fixed_row_height
                    ws.row_dimensions[current_row + 3].height = fixed_row_height
                    
                    # Apply borders to all data cells for this employee
                    for row in range(current_row, current_row + 4):
                        for col in range(1, current_col):
                            cell = ws.cell(row=row, column=col)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                    
                    # Apply alternating colors
                    for row in range(current_row, current_row + 4):
                        for col in range(1, current_col):
                            cell = ws.cell(row=row, column=col)
                            employee_group = (current_row - 5) // 4
                            if employee_group % 2 == 0:
                                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    
                    current_row += 4
                
                # Set column widths directly
                ws.column_dimensions['A'].width = 12  # Emp ID
                ws.column_dimensions['B'].width = 30  # Emp Name
                ws.column_dimensions['C'].width = 25  # Designation
                ws.column_dimensions['D'].width = 15   # Time
                
                # Set day column widths
                for col in range(5, current_col):
                    try:
                        column_letter = ws.cell(row=1, column=col).column_letter
                        ws.column_dimensions[column_letter].width = 12
                    except:
                        continue
                
                # Save workbook to ZIP
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                # Clean section name for filename
                clean_section_name = "".join(c for c in section if c.isalnum() or c in (' ', '-', '_')).rstrip()
                filename = f"{clean_section_name}_Attendance_Report.xlsx"
                
                zip_file.writestr(filename, excel_buffer.getvalue())
        
        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="staff_segregation_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
        
        return response
        
    except Exception as e:
        messages.error(request, f"Error generating segregation report: {str(e)}")
        return redirect('processor:file_detail', file_id=file_id)


@login_required(login_url='/app/login/')
def generate_detailed_attendance_report(request, file_id):
    """Generate detailed attendance report using template"""
    try:
        files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
        processed_file = get_object_or_404(files_queryset, id=file_id)
        
        # Get the processed file path
        if processed_file.processed_file:
            file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.processed_file))
        else:
            file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.original_file))
        
        if not os.path.exists(file_path):
            return JsonResponse({'error': 'File not found'}, status=404)
        
        # Load the processed data using matrix parser
        try:
            svc = ExcelProcessorService()
            df = svc.processor.process_matrix_attendance(file_path)
        except Exception:
            df = pd.read_excel(file_path)
        
        # Get unique employee IDs
        unique_employee_ids = df['Employee_ID'].dropna().unique()
        
        # Database connection
        conn = psycopg2.connect(
            host=config('DATABASE_HOST', default='localhost'),
            database=config('DATABASE_NAME', default='admin_db'),
            user=config('DATABASE_USER', default='postgres'),
            password=config('DATABASE_PASSWORD', default='Testing@123'),
            port=config('DATABASE_PORT', default='5432')
        )
        
        # Get staff details ordered by priority first, then employment type
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Build department filter
        department_filter = ""
        params = [list(unique_employee_ids)]
        if not request.user.is_superuser and request.user.department:
            department_filter = " AND department_id = %s"
            params.append(request.user.department.id)
        
        cursor.execute(f"""
            SELECT staffid, name, designation, level, section, weekly_off, type_of_employment, priority
            FROM staff_details 
            WHERE staffid = ANY(%s) 
            AND type_of_employment IN ('permanent', 'contract')
            {department_filter}
            ORDER BY 
                priority ASC,
                CASE 
                    WHEN type_of_employment = 'permanent' THEN 1
                    WHEN type_of_employment = 'contract' THEN 2
                    ELSE 3
                END,
                CASE 
                    WHEN REGEXP_REPLACE(staffid, '[^0-9]', '', 'g') ~ '^[0-9]+$' 
                    THEN CAST(REGEXP_REPLACE(staffid, '[^0-9]', '', 'g') AS INTEGER)
                    ELSE 999999
                END ASC
        """, params)
        
        staff_details = {row['staffid']: row for row in cursor.fetchall()}
        cursor.close()
        conn.close()
        
        # Load the template
        template_path = os.path.join(settings.BASE_DIR, 'static', 'detailed_attendance_template.xlsx')
        
        if not os.path.exists(template_path):
            return JsonResponse({'error': 'Template file not found'}, status=404)
        
        # Load template and create a copy
        wb = openpyxl.load_workbook(template_path)
        ws = wb['Template Sheet']  # Use the specific sheet name
        
        # Copy row heights from template
        for row_num in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_num].height is not None:
                # Preserve the original row height
                original_height = ws.row_dimensions[row_num].height
                ws.row_dimensions[row_num].height = original_height
        
        # Extract and normalize period from the original file
        period = "Unknown"
        try:
            original_file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.original_file))
            if os.path.exists(original_file_path):
                original_df = pd.read_excel(original_file_path, header=None)
                if len(original_df) >= 9:
                    period_cell = original_df.iloc[8, 0]
                    if pd.notna(period_cell):
                        import re
                        raw_text = str(period_cell).strip()
                        # Remove leading 'Period:' (any case) and surrounding spaces
                        period = re.sub(r"(?i)^\s*period\s*:?,?\s*", "", raw_text) or "Unknown"
        except Exception as e:
            logger.error(f"Error extracting period: {e}")
        
        # Helper function to safely set cell value
        def safe_set_cell(worksheet, cell_address, value):
            """Safely set cell value, handling merged cells"""
            try:
                cell = worksheet[cell_address]
                # Check if cell is part of a merged range
                for merged_range in worksheet.merged_cells.ranges:
                    if cell_address in merged_range:
                        # Set value to the top-left cell of the merged range
                        top_left_cell = merged_range.top_left
                        worksheet[top_left_cell] = value
                        return
                # If not merged, set normally
                worksheet[cell_address] = value
            except Exception as e:
                logger.warning(f"Could not set cell {cell_address}: {e}")
        
        # Fill period in E1 (single 'Period:' prefix)
        safe_set_cell(ws, 'E1', f"Period: {period}")
        
        # Calculate total number of days in the selected period
        total_days = 0
        try:
            # Extract start and end dates from period string
            if ' - ' in period:
                period_parts = period.split(' - ')
                if len(period_parts) == 2:
                    start_date_str = period_parts[0].strip()
                    end_date_str = period_parts[1].strip()
                    
                    # Parse the dates (assuming format like 2082/03/01)
                    try:
                        # Split by '/' and extract year, month, day
                        start_parts = start_date_str.split('/')
                        end_parts = end_date_str.split('/')
                        
                        if len(start_parts) == 3 and len(end_parts) == 3:
                            start_day = int(start_parts[2])
                            end_day = int(end_parts[2])
                            
                            # Calculate total days (inclusive)
                            total_days = end_day - start_day + 1
                        else:
                            # Fallback: count unique dates from data
                            unique_dates = df['Date'].dropna().unique()
                            total_days = len(unique_dates)
                    except (ValueError, IndexError):
                        # Fallback: count unique dates from data
                        unique_dates = df['Date'].dropna().unique()
                        total_days = len(unique_dates)
                else:
                    # Fallback: count unique dates from data
                    unique_dates = df['Date'].dropna().unique()
                    total_days = len(unique_dates)
            else:
                # Fallback: count unique dates from data
                unique_dates = df['Date'].dropna().unique()
                total_days = len(unique_dates)
        except Exception as e:
            logger.error(f"Error calculating total days: {e}")
            total_days = 0
        
        # Fill total days in F2
        safe_set_cell(ws, 'F2', total_days)
        
        # Process each employee in sorted order
        row = 4  # Start from row 4
        sorted_staff = list(staff_details.values())
        for staff in sorted_staff:
            emp_id = staff['staffid']
            emp_data = df[df['Employee_ID'] == emp_id]
            
            # Calculate attendance statistics
            present_days = 0
            absent_days = 0
            weekly_off_days = 0
            allowance_days = 0
            personal_leave_days = 0
            sick_leave_days = 0
            casual_leave_days = 0
            substitute_leave_days = 0
            duty_leave_days = 0
            other_leave_days = 0
            
            present_dates = []
            absent_dates = []
            weekly_off_dates = []
            allowance_dates = []
            personal_leave_dates = []
            sick_leave_dates = []
            casual_leave_dates = []
            substitute_leave_dates = []
            duty_leave_dates = []
            other_leave_dates = []
            
            # Process each day for this employee
            for _, emp_row in emp_data.iterrows():
                try:
                    status = str(emp_row.get('Status', '')).strip().upper()
                    date = emp_row.get('Date', '')
                    in_time = emp_row.get('InTime', '')
                    out_time = emp_row.get('OutTime', '')
                except (KeyError, AttributeError):
                    continue
                
                # Determine attendance status
                if 'P' in status or 'A *' in status:
                    present_days += 1
                    present_dates.append(date)
                    allowance_days += 1
                    allowance_dates.append(date)
                elif status == 'A':
                    absent_days += 1
                    absent_dates.append(date)
                elif 'WO' in status or 'HO' in status:
                    present_days += 1
                    present_dates.append(date)
                    weekly_off_days += 1
                    weekly_off_dates.append(date)
                    
                    # Check if there's actual work time recorded for allowance
                    has_work_time = (in_time and str(in_time).strip() != '' and str(in_time).strip() != 'nan') or \
                                  (out_time and str(out_time).strip() != '' and str(out_time).strip() != 'nan')
                    if has_work_time:
                        allowance_days += 1
                        allowance_dates.append(date)
                elif 'PL' in status:
                    personal_leave_days += 1
                    personal_leave_dates.append(date)
                elif 'SL' in status:
                    sick_leave_days += 1
                    sick_leave_dates.append(date)
                elif 'CL' in status:
                    casual_leave_days += 1
                    casual_leave_dates.append(date)
                elif 'SUBSTITUTE' in status or 'SUBL' in status:
                    substitute_leave_days += 1
                    substitute_leave_dates.append(date)
                elif 'DUTY' in status:
                    duty_leave_days += 1
                    duty_leave_dates.append(date)
                    # Also count as other leave for the M column (but don't add to other_leave_dates to avoid duplication in remarks)
                    other_leave_days += 1
                elif 'L' in status:
                    # Count as other leave for the M column
                    other_leave_days += 1
                    other_leave_dates.append(date)
                else:
                    # Count as other leave if status doesn't match any specific leave type (PL, SL, CL, SUBSTITUTE)
                    other_leave_days += 1
                    other_leave_dates.append(date)
            
            # Fill data in the template
            safe_set_cell(ws, f'B{row}', staff['name'].title())  # Proper case name
            safe_set_cell(ws, f'C{row}', staff['staffid'])
            safe_set_cell(ws, f'D{row}', staff['designation'])
            safe_set_cell(ws, f'E{row}', staff['level'])
            safe_set_cell(ws, f'F{row}', present_days)
            safe_set_cell(ws, f'G{row}', personal_leave_days)  # PL count
            safe_set_cell(ws, f'H{row}', sick_leave_days)      # SL count
            safe_set_cell(ws, f'I{row}', casual_leave_days)    # CL count
            safe_set_cell(ws, f'J{row}', substitute_leave_days) # Substitute count
            safe_set_cell(ws, f'L{row}', absent_days)          # Absent count
            safe_set_cell(ws, f'M{row}', other_leave_days)     # Other leave count
            safe_set_cell(ws, f'N{row}', allowance_days)       # Allowance count
            
            # Fill weekly off day name instead of count
            weekly_off_day = staff.get('weekly_off', '').title() if staff.get('weekly_off') else ''
            safe_set_cell(ws, f'R{row}', weekly_off_day)
            
            # Create leave details string (show only day numbers)
            import re
            def day_only(val):
                s = str(val)
                m = re.match(r"\s*(\d{1,2})\b", s)
                return m.group(1) if m else s
            def join_days(values):
                return ", ".join(day_only(v) for v in values)
            leave_details = []
            if personal_leave_dates:
                leave_details.append(f"PL on {join_days(personal_leave_dates)}")
            if casual_leave_dates:
                leave_details.append(f"CL on {join_days(casual_leave_dates)}")
            if sick_leave_dates:
                leave_details.append(f"SL on {join_days(sick_leave_dates)}")
            if substitute_leave_dates:
                leave_details.append(f"SUBSTITUTE on {join_days(substitute_leave_dates)}")
            if duty_leave_dates:
                leave_details.append(f"DUTY on {join_days(duty_leave_dates)}")
            if other_leave_dates:
                leave_details.append(f"Other on {join_days(other_leave_dates)}")
            if absent_dates:
                leave_details.append(f"Absent on {join_days(absent_dates)}")
            
            safe_set_cell(ws, f'S{row}', ', '.join(leave_details))
            
            row += 1
        
        # Save the filled template
        output_filename = f"detailed_attendance_{processed_file.id}.xlsx"
        output_path = os.path.join(settings.MEDIA_ROOT, 'processed', output_filename)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb.save(output_path)
        
        # Return the file for download
        with open(output_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            return response
            
    except Exception as e:
        logger.error(f"Error generating detailed attendance report: {str(e)}")
        return JsonResponse({'error': f'Error generating detailed attendance report: {str(e)}'}, status=500)


@login_required(login_url='/app/login/')
def generate_monthly_wages_report(request, file_id):
    """Generate monthly wages report using template"""
    try:
        files_queryset = get_department_filtered_queryset(request.user, ProcessedFile)
        processed_file = get_object_or_404(files_queryset, id=file_id)
        
        # Get the processed file path
        if processed_file.processed_file:
            file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.processed_file))
        else:
            file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.original_file))
        
        if not os.path.exists(file_path):
            return JsonResponse({'error': 'File not found'}, status=404)
        
        # Load the processed data using matrix parser
        try:
            svc = ExcelProcessorService()
            df = svc.processor.process_matrix_attendance(file_path)
        except Exception:
            df = pd.read_excel(file_path)
        
        # Get unique employee IDs
        unique_employee_ids = df['Employee_ID'].dropna().unique()
        
        # Database connection
        conn = psycopg2.connect(
            host=config('DATABASE_HOST', default='localhost'),
            database=config('DATABASE_NAME', default='admin_db'),
            user=config('DATABASE_USER', default='postgres'),
            password=config('DATABASE_PASSWORD', default='Testing@123'),
            port=config('DATABASE_PORT', default='5432')
        )
        
        # Get staff details ordered by priority first, then employment type
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Build department filter
        department_filter = ""
        params = [list(unique_employee_ids)]
        if not request.user.is_superuser and request.user.department:
            department_filter = " AND department_id = %s"
            params.append(request.user.department.id)
        
        cursor.execute(f"""
            SELECT staffid, name, designation, level, section, weekly_off, type_of_employment, priority
            FROM staff_details 
            WHERE staffid = ANY(%s) 
            AND type_of_employment = 'monthly wages'
            {department_filter}
            ORDER BY 
                priority ASC,
                CASE 
                    WHEN REGEXP_REPLACE(staffid, '[^0-9]', '', 'g') ~ '^[0-9]+$' 
                    THEN CAST(REGEXP_REPLACE(staffid, '[^0-9]', '', 'g') AS INTEGER)
                    ELSE 999999
                END ASC
        """, params)
        
        staff_details = {row['staffid']: row for row in cursor.fetchall()}
        cursor.close()
        conn.close()
        
        # Load the template (using the same template for now)
        template_path = os.path.join(settings.BASE_DIR, 'static', 'detailed_attendance_template.xlsx')
        
        if not os.path.exists(template_path):
            return JsonResponse({'error': 'Template file not found'}, status=404)
        
        # Load template and create a copy
        wb = openpyxl.load_workbook(template_path)
        ws = wb['Template Sheet']  # Use the specific sheet name
        
        # Copy row heights from template
        for row_num in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_num].height is not None:
                # Preserve the original row height
                original_height = ws.row_dimensions[row_num].height
                ws.row_dimensions[row_num].height = original_height
        
        # Extract period from the original file
        period = "Unknown Period"
        try:
            original_file_path = os.path.join(settings.MEDIA_ROOT, str(processed_file.original_file))
            if os.path.exists(original_file_path):
                # Read the original file to get period
                original_df = pd.read_excel(original_file_path, header=None)
                if len(original_df) >= 9:
                    period_cell = original_df.iloc[8, 0]  # A9 cell
                    if pd.notna(period_cell):
                        period = str(period_cell)
        except Exception as e:
            logger.error(f"Error extracting period: {e}")
        
        # Helper function to safely set cell value
        def safe_set_cell(worksheet, cell_address, value):
            """Safely set cell value, handling merged cells"""
            try:
                cell = worksheet[cell_address]
                # Check if cell is part of a merged range
                for merged_range in worksheet.merged_cells.ranges:
                    if cell_address in merged_range:
                        # Set value to the top-left cell of the merged range
                        top_left_cell = merged_range.top_left
                        worksheet[top_left_cell] = value
                        return
                # If not merged, set normally
                worksheet[cell_address] = value
            except Exception as e:
                logger.warning(f"Could not set cell {cell_address}: {e}")
        
        # Fill period in E1
        safe_set_cell(ws, 'E1', f"Period: {period}")
        
        # Calculate total number of days in the selected period
        total_days = 0
        try:
            # Extract start and end dates from period string
            if ' - ' in period:
                period_parts = period.split(' - ')
                if len(period_parts) == 2:
                    start_date_str = period_parts[0].strip()
                    end_date_str = period_parts[1].strip()
                    
                    # Parse the dates (assuming format like 2082/03/01)
                    try:
                        # Split by '/' and extract year, month, day
                        start_parts = start_date_str.split('/')
                        end_parts = end_date_str.split('/')
                        
                        if len(start_parts) == 3 and len(end_parts) == 3:
                            start_day = int(start_parts[2])
                            end_day = int(end_parts[2])
                            
                            # Calculate total days (inclusive)
                            total_days = end_day - start_day + 1
                        else:
                            # Fallback: count unique dates from data
                            unique_dates = df['Date'].dropna().unique()
                            total_days = len(unique_dates)
                    except (ValueError, IndexError):
                        # Fallback: count unique dates from data
                        unique_dates = df['Date'].dropna().unique()
                        total_days = len(unique_dates)
                else:
                    # Fallback: count unique dates from data
                    unique_dates = df['Date'].dropna().unique()
                    total_days = len(unique_dates)
            else:
                # Fallback: count unique dates from data
                unique_dates = df['Date'].dropna().unique()
                total_days = len(unique_dates)
        except Exception as e:
            logger.error(f"Error calculating total days: {e}")
            total_days = 0
        
        # Fill total days in F2
        safe_set_cell(ws, 'F2', total_days)
        
        # Process each employee in sorted order
        row = 4  # Start from row 4
        sorted_staff = list(staff_details.values())
        for staff in sorted_staff:
            emp_id = staff['staffid']
            emp_data = df[df['Employee_ID'] == emp_id]
            
            # Calculate attendance statistics
            present_days = 0
            absent_days = 0
            weekly_off_days = 0
            allowance_days = 0
            personal_leave_days = 0
            sick_leave_days = 0
            casual_leave_days = 0
            substitute_leave_days = 0
            duty_leave_days = 0
            other_leave_days = 0
            
            present_dates = []
            absent_dates = []
            weekly_off_dates = []
            allowance_dates = []
            personal_leave_dates = []
            sick_leave_dates = []
            casual_leave_dates = []
            substitute_leave_dates = []
            duty_leave_dates = []
            other_leave_dates = []
            
            # Process each day for this employee
            for _, emp_row in emp_data.iterrows():
                try:
                    status = str(emp_row.get('Status', '')).strip().upper()
                    date = emp_row.get('Date', '')
                    in_time = emp_row.get('InTime', '')
                    out_time = emp_row.get('OutTime', '')
                except (KeyError, AttributeError):
                    continue
                
                # Determine attendance status
                if 'P' in status or 'A *' in status:
                    present_days += 1
                    present_dates.append(date)
                    allowance_days += 1
                    allowance_dates.append(date)
                elif status == 'A':
                    absent_days += 1
                    absent_dates.append(date)
                elif 'WO' in status or 'HO' in status:
                    present_days += 1
                    present_dates.append(date)
                    weekly_off_days += 1
                    weekly_off_dates.append(date)
                    
                    # Check if there's actual work time recorded for allowance
                    has_work_time = (in_time and str(in_time).strip() != '' and str(in_time).strip() != 'nan') or \
                                  (out_time and str(out_time).strip() != '' and str(out_time).strip() != 'nan')
                    if has_work_time:
                        allowance_days += 1
                        allowance_dates.append(date)
                elif 'PL' in status:
                    personal_leave_days += 1
                    personal_leave_dates.append(date)
                elif 'SL' in status:
                    sick_leave_days += 1
                    sick_leave_dates.append(date)
                elif 'CL' in status:
                    casual_leave_days += 1
                    casual_leave_dates.append(date)
                elif 'SUBSTITUTE' in status or 'SUBL' in status:
                    substitute_leave_days += 1
                    substitute_leave_dates.append(date)
                elif 'DUTY' in status:
                    duty_leave_days += 1
                    duty_leave_dates.append(date)
                    # Also count as other leave for the M column (but don't add to other_leave_dates to avoid duplication in remarks)
                    other_leave_days += 1
                elif 'L' in status:
                    # Count as other leave for the M column
                    other_leave_days += 1
                    other_leave_dates.append(date)
                else:
                    # Count as other leave if status doesn't match any specific leave type (PL, SL, CL, SUBSTITUTE)
                    other_leave_days += 1
                    other_leave_dates.append(date)
            
            # Fill data in the template
            safe_set_cell(ws, f'B{row}', staff['name'].title())  # Proper case name
            safe_set_cell(ws, f'C{row}', staff['staffid'])
            safe_set_cell(ws, f'D{row}', staff['designation'])
            safe_set_cell(ws, f'E{row}', staff['level'])
            safe_set_cell(ws, f'F{row}', present_days)
            safe_set_cell(ws, f'G{row}', personal_leave_days)  # PL count
            safe_set_cell(ws, f'H{row}', sick_leave_days)      # SL count
            safe_set_cell(ws, f'I{row}', casual_leave_days)    # CL count
            safe_set_cell(ws, f'J{row}', substitute_leave_days) # Substitute count
            safe_set_cell(ws, f'L{row}', absent_days)          # Absent count
            safe_set_cell(ws, f'M{row}', other_leave_days)     # Other leave count
            safe_set_cell(ws, f'N{row}', allowance_days)       # Allowance count
            
            # Fill weekly off day name instead of count
            weekly_off_day = staff.get('weekly_off', '').title() if staff.get('weekly_off') else ''
            safe_set_cell(ws, f'R{row}', weekly_off_day)
            
            # Create leave details string (show only day numbers)
            import re
            def day_only(val):
                s = str(val)
                m = re.match(r"\s*(\d{1,2})\b", s)
                return m.group(1) if m else s
            def join_days(values):
                return ", ".join(day_only(v) for v in values)
            leave_details = []
            if personal_leave_dates:
                leave_details.append(f"PL on {join_days(personal_leave_dates)}")
            if casual_leave_dates:
                leave_details.append(f"CL on {join_days(casual_leave_dates)}")
            if sick_leave_dates:
                leave_details.append(f"SL on {join_days(sick_leave_dates)}")
            if substitute_leave_dates:
                leave_details.append(f"SUBSTITUTE on {join_days(substitute_leave_dates)}")
            if duty_leave_dates:
                leave_details.append(f"DUTY on {join_days(duty_leave_dates)}")
            if other_leave_dates:
                leave_details.append(f"Other on {join_days(other_leave_dates)}")
            if absent_dates:
                leave_details.append(f"Absent on {join_days(absent_dates)}")
            
            safe_set_cell(ws, f'S{row}', ', '.join(leave_details))
            
            row += 1
        
        # Save the filled template
        output_filename = f"monthly_wages_attendance_{processed_file.id}.xlsx"
        output_path = os.path.join(settings.MEDIA_ROOT, 'processed', output_filename)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb.save(output_path)
        
        # Return the file for download
        with open(output_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{output_filename}"'
            return response
            
    except Exception as e:
        logger.error(f"Error generating monthly wages report: {str(e)}")
        return JsonResponse({'error': f'Error generating monthly wages report: {str(e)}'}, status=500)
