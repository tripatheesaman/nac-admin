"""
Service layer for Excel processing following SOLID principles
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Any, Optional, Callable
import logging
from abc import ABC, abstractmethod

logger = logging.getLogger(__name__)


class ExcelProcessorInterface(ABC):
    """Interface for Excel processing following Interface Segregation Principle"""
    
    @abstractmethod
    def read_excel(self, file_path: str) -> pd.DataFrame:
        """Read Excel file and return DataFrame"""
        pass
    
    @abstractmethod
    def process_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """Process the data according to business rules"""
        pass
    
    @abstractmethod
    def write_excel(self, data: pd.DataFrame, output_path: str) -> None:
        """Write processed data to Excel file"""
        pass


class ExcelReader:
    """Single Responsibility: Read Excel files"""
    
    def read_excel(self, file_path: str, progress_callback: Optional[Callable] = None) -> pd.DataFrame:
        """Read Excel file and return DataFrame"""
        try:
            if progress_callback:
                progress_callback(20, "Reading Excel file...")
            
            # Try different engines based on file extension
            if file_path.lower().endswith('.xls'):
                df = pd.read_excel(file_path, engine='xlrd')
            else:
                df = pd.read_excel(file_path, engine='openpyxl')
            
            if progress_callback:
                progress_callback(30, f"Successfully loaded {len(df)} rows from Excel file")
            
            logger.info(f"Successfully read Excel file: {file_path}")
            return df
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            raise ValueError(f"Could not read Excel file: {str(e)}")


class DataProcessor:
    """Single Responsibility: Process data according to business rules"""
    
    def process_data(self, data: pd.DataFrame, progress_callback: Optional[Callable] = None) -> pd.DataFrame:
        """Process the attendance data from Report (1).xls format"""
        try:
            if progress_callback:
                progress_callback(40, "Processing attendance data...")
            
            logger.info("Starting attendance data processing")
            
            # Process the actual data
            if data.empty:
                raise ValueError("Input data is empty")

            # Assuming standard column names from the input file
            required_columns = ['Employee ID', 'Employee Name', 'Designation', 'Date', 'In Time', 'Out Time']
            missing_columns = [col for col in required_columns if col not in data.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

            # Create processed DataFrame
            processed_data = pd.DataFrame()
            processed_data['Employee_ID'] = data['Employee ID']
            processed_data['Employee_Name'] = data['Employee Name']
            processed_data['Designation'] = data['Designation']
            processed_data['Date'] = pd.to_datetime(data['Date']).dt.date
            processed_data['Day_Name'] = pd.to_datetime(data['Date']).dt.day_name()
            processed_data['InTime'] = pd.to_datetime(data['In Time']).dt.strftime('%H:%M')
            processed_data['OutTime'] = pd.to_datetime(data['Out Time']).dt.strftime('%H:%M')

            # Calculate worked hours
            in_time = pd.to_datetime(data['In Time'])
            out_time = pd.to_datetime(data['Out Time'])
            worked_hours = (out_time - in_time).dt.total_seconds() / 3600
            processed_data['WorkedHours'] = worked_hours.round(2)

            # Determine status based on worked hours
            processed_data['Status'] = processed_data['WorkedHours'].apply(
                lambda x: 'Present' if x >= 8 else 'Half Day' if x >= 4 else 'Absent'
            )
            
            if progress_callback:
                progress_callback(60, "Data processing completed")
            
            logger.info("Data processing completed successfully")
            return processed_data
            
        except Exception as e:
            logger.error(f"Error processing data: {str(e)}")
            raise ValueError(f"Data processing failed: {str(e)}")
    
    def process_attendance_file(self, file_path: str, progress_callback: Optional[Callable] = None) -> pd.DataFrame:
        """Process the specific Report (1).xls attendance file format"""
        try:
            import pandas as pd
            from datetime import datetime
            import re
            
            if progress_callback:
                progress_callback(40, "Processing attendance file...")
            
            logger.info(f"Processing attendance file: {file_path}")
            
            # Read the Excel file using pandas with appropriate engine
            if file_path.lower().endswith('.xls'):
                df = pd.read_excel(file_path, engine='xlrd', header=None)
            else:
                df = pd.read_excel(file_path, engine='openpyxl', header=None)
            
            if progress_callback:
                progress_callback(50, f"Loaded {len(df)} rows, extracting data...")
            
            logger.info(f"Successfully loaded Excel file with {len(df)} rows")
            
            # Extract period from cell A9 (row 8, column 0 in pandas)
            period_cell = df.iloc[8, 0] if len(df) > 8 else None
            logger.info(f"Period cell content: {period_cell}")
            
            # Helper to detect if a header cell is date-like
            def is_date_like(value) -> bool:
                try:
                    if value is None:
                        return False
                    # Datetime-like objects
                    from datetime import datetime, date
                    if isinstance(value, (datetime, date)):
                        return True
                    # Excel serial date (rough bounds)
                    if isinstance(value, (int, float)) and 20000 <= float(value) <= 50000:
                        return True
                    # Strings containing plausible date patterns
                    text = str(value).strip()
                    if not text:
                        return False
                    if re.search(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text):
                        return True
                    # Year-month-day or day-month with names could be added here if needed
                    return False
                except Exception:
                    return False

            # Locate header row (default row 10). If it fails to yield date columns, scan a window to find a better header row.
            def find_header_row(frame) -> int:
                default_idx = 10
                candidate_indices = list(range(default_idx, min(default_idx + 10, len(frame))))
                for idx in candidate_indices:
                    row = frame.iloc[idx]
                    # Count date-like cells from column 5 onwards
                    date_like_count = 0
                    for i in range(5, len(row)):
                        if is_date_like(row.iloc[i]):
                            date_like_count += 1
                        # Early accept if enough
                        if date_like_count >= 3:
                            return idx
                return default_idx

            header_row_index = find_header_row(df)
            data_start_row = header_row_index
            if len(df) <= data_start_row:
                raise ValueError("File does not contain enough data rows")

            headers = df.iloc[data_start_row]

            # Find the date columns (starting from column F, index 5)
            date_columns = []
            for i in range(5, len(headers)):
                cell_value = headers.iloc[i]
                if is_date_like(cell_value):
                    date_columns.append(i)
            
            if progress_callback:
                progress_callback(60, f"Found {len(date_columns)} date columns, processing attendance...")

            # If no date columns detected, fail fast with a helpful error instead of returning 0 rows
            if len(date_columns) == 0:
                logger.error("Attendance parsing error: No date columns detected in header row %s", data_start_row)
                raise ValueError("Could not detect date columns. Please verify the file layout or share a sample for calibration.")
            
            logger.info(f"Found {len(date_columns)} date columns")
            
            # Process each employee row
            processed_data = []
            
            logger.info(f"Starting to process employee rows from row {data_start_row + 1}")
            
            # Start from the row after headers for employee data
            for row_idx in range(data_start_row + 1, len(df)):
                row = df.iloc[row_idx]
                
                # Skip empty rows
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == '':
                    logger.debug(f"Skipping empty row {row_idx}")
                    continue
                
                logger.info(f"Processing row {row_idx}: Employee ID: {str(row.iloc[0]).strip()}")
                
                # Extract employee information
                employee_id = str(row.iloc[0]).strip()
                employee_name = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ''
                designation = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ''
                
                # Process each date column
                logger.info(f"Found {len(date_columns)} date columns to process for this employee")
                for col_idx in date_columns:
                    date_header = headers.iloc[col_idx]
                    attendance_value = row.iloc[col_idx]
                    
                    logger.debug(f"Processing column {col_idx}, date: {date_header}, value: {attendance_value}")
                    
                    if pd.isna(attendance_value) or str(attendance_value).strip() == '':
                        logger.debug(f"Skipping empty attendance value for column {col_idx}")
                        continue
                    
                    # Parse the attendance value
                    attendance_str = str(attendance_value).strip()
                    
                    # Extract time information
                    in_time = None
                    out_time = None
                    status = 'Present'
                    worked_hours = 0
                    
                    if 'P' in attendance_str or 'Present' in attendance_str:
                        status = 'Present'
                        # Extract both in and out times
                        time_matches = re.findall(r'(\d{1,2}:\d{2})', attendance_str)
                        if len(time_matches) >= 2:
                            in_time = time_matches[0]
                            out_time = time_matches[1]
                            try:
                                # Validate times
                                datetime.strptime(in_time, '%H:%M')
                                datetime.strptime(out_time, '%H:%M')
                            except ValueError:
                                in_time = None
                                out_time = None
                        elif len(time_matches) == 1:
                            in_time = time_matches[0]
                    elif 'A' in attendance_str or 'Absent' in attendance_str:
                        status = 'Absent'
                    elif 'L' in attendance_str or 'Leave' in attendance_str:
                        status = 'Leave'
                    elif 'H' in attendance_str or 'Holiday' in attendance_str:
                        status = 'Holiday'
                        
                    # Add logging to debug time extraction
                    logger.debug(f"Attendance string: {attendance_str}")
                    logger.debug(f"Extracted times: in={in_time}, out={out_time}")
                    
                    # Calculate worked hours if both in and out times are available
                    if in_time and out_time:
                        try:
                            from datetime import datetime, timedelta
                            in_dt = datetime.strptime(in_time, '%H:%M')
                            out_dt = datetime.strptime(out_time, '%H:%M')
                            if out_dt < in_dt:
                                out_dt += timedelta(days=1)
                            worked_hours = (out_dt - in_dt).total_seconds() / 3600
                        except:
                            worked_hours = 0
                    
                    # Parse date from header robustly
                    try:
                        if isinstance(date_header, str):
                            text = date_header.strip()
                            # Try common formats
                            fmt_candidates = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']
                            parsed = None
                            for fmt in fmt_candidates:
                                try:
                                    parsed = datetime.strptime(text, fmt)
                                    break
                                except Exception:
                                    continue
                            if parsed is None:
                                continue
                            date_obj = parsed
                        elif isinstance(date_header, (int, float)):
                            # Excel serial date
                            from datetime import datetime, timedelta
                            excel_base = datetime(1899, 12, 30)  # Excel's epoch
                            date_obj = excel_base + timedelta(days=float(date_header))
                        else:
                            date_obj = date_header

                        day_name = date_obj.strftime('%A')
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except Exception:
                        continue
                    
                    processed_data.append({
                        'Employee_ID': employee_id,
                        'Employee_Name': employee_name,
                        'Designation': designation,
                        'Date': date_str,
                        'Day_Name': day_name,
                        'InTime': in_time,
                        'OutTime': out_time,
                        'Status': status,
                        'WorkedHours': worked_hours
                    })
            
            if progress_callback:
                progress_callback(70, f"Processed {len(processed_data)} attendance records")
            
            logger.info(f"Successfully processed {len(processed_data)} attendance records")
            
            df_out = pd.DataFrame(processed_data)
            if df_out.empty:
                raise ValueError("Parsed attendance produced 0 rows. Verify header row and date columns.")
            return df_out
            
        except Exception as e:
            logger.error(f"Error processing attendance file: {str(e)}")
            raise ValueError(f"Attendance file processing failed: {str(e)}")

    def process_matrix_attendance(self, file_path: str, progress_callback: Optional[Callable] = None) -> pd.DataFrame:
        """Process matrix-style attendance where header has fixed ID/Name/Designation and day columns like '1 Mon'."""
        import re
        from datetime import datetime, timedelta
        try:
            if progress_callback:
                progress_callback(40, "Processing matrix attendance file...")

            # Read sheet without headers to locate header row dynamically
            if file_path.lower().endswith('.xls'):
                raw = pd.read_excel(file_path, engine='xlrd', header=None)
            else:
                raw = pd.read_excel(file_path, engine='openpyxl', header=None)

            # Find header row by detecting multiple day-like headers in a row
            header_row_idx = None
            day_header_pattern = re.compile(r"^\s*\d{1,2}\s+[A-Za-z]+\s*$")
            scan_limit = min(50, len(raw))
            for r in range(scan_limit):
                row_vals = raw.iloc[r].tolist()
                matches = [i for i, v in enumerate(row_vals) if isinstance(v, str) and day_header_pattern.match(v.strip())]
                # Only consider day columns if they occur after the first 3 columns
                matches = [i for i in matches if i >= 3]
                if len(matches) >= 3:
                    header_row_idx = r
                    day_col_indices = matches
                    break

            if header_row_idx is None:
                raise ValueError("Could not locate header row with day headers like '1 Mon'")

            headers = raw.iloc[header_row_idx].tolist()
            # Identify important columns by header labels (case-insensitive)
            def find_col(label_set, default_idx=None):
                label_set = {s.lower() for s in label_set}
                for idx, val in enumerate(headers):
                    text = str(val).strip().lower() if val is not None else ''
                    if text in label_set:
                        return idx
                return default_idx

            sn_col = find_col({"sn", "sn.", "s.n."}, 0)
            emp_id_col = find_col({"emp id", "empid", "employee id", "emp no", "emp no."}, 1)
            name_col = find_col({"name", "employee name", "emp name"}, 2)
            post_col = find_col({"post", "designation", "desig", "designation"}, 3)
            time_col = find_col({"time"}, 4)

            # Confirm or compute day_col_indices from headers (to the right of the Time column)
            if not day_col_indices:
                day_col_indices = []
                for idx, val in enumerate(headers):
                    if idx <= (time_col if time_col is not None else 3):
                        continue
                    text = str(val).strip() if val is not None else ''
                    if day_header_pattern.match(text):
                        day_col_indices.append(idx)

            if not day_col_indices:
                raise ValueError("No day columns found (expected headers like '1 Mon', '2 Tue')")

            processed_rows = []

            # Helpers
            def cell_str(val):
                return '' if pd.isna(val) else str(val).strip()

            # Accumulators for a single employee block
            last_emp_id = ''
            last_name = ''
            last_post = ''
            cur_in = None
            cur_out = None
            cur_status = None
            cur_hours = None

            def extract_day_values(row):
                return [cell_str(row.iloc[i]) for i in day_col_indices]

            def flush_block():
                nonlocal cur_in, cur_out, cur_status, cur_hours, last_emp_id, last_name, last_post
                if cur_in is None and cur_out is None and cur_status is None and cur_hours is None:
                    return
                total_days = len(day_col_indices)
                for j in range(total_days):
                    raw_header = headers[day_col_indices[j]] if j < total_days else ''
                    day_header = str(raw_header).strip()
                    # Normalize header and derive parts; ensure single weekday token
                    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)$", day_header)
                    if m:
                        day_num = m.group(1).zfill(2)
                        day_name = m.group(2).strip()
                        date_text = f"{day_num} {day_name}"
                    else:
                        # Fallback: take first two tokens if available
                        parts = day_header.split()
                        if len(parts) >= 2 and parts[0].isdigit():
                            day_num = parts[0].zfill(2)
                            day_name = parts[1]
                            date_text = f"{day_num} {day_name}"
                        else:
                            day_name = day_header
                            date_text = day_header or ''
                    in_time = cur_in[j] if cur_in and j < len(cur_in) else ''
                    out_time = cur_out[j] if cur_out and j < len(cur_out) else ''
                    status_val = cur_status[j] if cur_status and j < len(cur_status) else ''
                    hours_val = cur_hours[j] if cur_hours and j < len(cur_hours) else ''
                    # Skip if all empty
                    if not any([in_time, out_time, status_val, hours_val]):
                        continue
                    processed_rows.append({
                        'Employee_ID': last_emp_id,
                        'Employee_Name': last_name,
                        'Designation': last_post,
                        'Date': date_text,
                        'Day_Name': day_name,
                        'InTime': in_time,
                        'OutTime': out_time,
                        'Status': status_val,
                        'WorkedHours': hours_val,
                    })
                cur_in = cur_out = cur_status = cur_hours = None

            # Iterate rows
            for r in range(header_row_idx + 1, len(raw)):
                row = raw.iloc[r]

                # Update identity if present on this row
                id_val = cell_str(row.iloc[emp_id_col]) if emp_id_col is not None and emp_id_col < len(row) else ''
                name_val = cell_str(row.iloc[name_col]) if name_col is not None and name_col < len(row) else ''
                post_val = cell_str(row.iloc[post_col]) if post_col is not None and post_col < len(row) else ''

                if id_val:
                    last_emp_id = id_val
                if name_val:
                    last_name = name_val
                if post_val:
                    last_post = post_val

                # Row type by 'Time' column label
                label = cell_str(row.iloc[time_col]) if time_col is not None and time_col < len(row) else ''
                label_low = label.lower()
                if label_low.startswith('intime') or label_low == 'in time':
                    # If a new InTime begins while we already have data, flush previous block first
                    if any([cur_in, cur_out, cur_status, cur_hours]):
                        flush_block()
                    cur_in = extract_day_values(row)
                    continue
                if label_low.startswith('out'):
                    cur_out = extract_day_values(row)
                    continue
                if 'status' in label_low:
                    cur_status = [cell_str(v).upper() for v in extract_day_values(row)]
                    continue
                if 'work' in label_low:
                    cur_hours = extract_day_values(row)
                    # We assume a block ends after worked hours row
                    flush_block()
                    continue

            # Flush any remaining at EOF
            flush_block()

            df_out = pd.DataFrame(processed_rows)
            if df_out.empty:
                raise ValueError("Parsed matrix attendance produced 0 rows. Verify header row and day columns.")

            if progress_callback:
                progress_callback(70, f"Processed {len(df_out)} attendance records")

            return df_out
        except Exception as e:
            logger.error(f"Error processing matrix attendance file: {str(e)}")
            raise ValueError(f"Matrix attendance processing failed: {str(e)}")


class ExcelWriter:
    """Single Responsibility: Write Excel files"""
    
    def write_excel(self, data: pd.DataFrame, output_path: str, template_path: Optional[str] = None, progress_callback: Optional[Callable] = None) -> None:
        """Write processed data to Excel file"""
        try:
            if progress_callback:
                progress_callback(80, "Writing output file...")
            
            if template_path:
                self._write_with_template(data, output_path, template_path)
            else:
                self._write_new_file(data, output_path)
            
            if progress_callback:
                progress_callback(90, "Finalizing output file...")
            
            logger.info(f"Successfully wrote Excel file: {output_path}")
            
        except Exception as e:
            logger.error(f"Error writing Excel file {output_path}: {str(e)}")
            raise ValueError(f"Could not write Excel file: {str(e)}")
    
    def _write_with_template(self, data: pd.DataFrame, output_path: str, template_path: str) -> None:
        """Write data using an existing template"""
        # Load template
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb.active
        
        # Write data to template
        for idx, row in data.iterrows():
            for col_idx, value in enumerate(row, 1):
                template_ws.cell(row=idx + 2, column=col_idx, value=value)
        
        # Save
        template_wb.save(output_path)
    
    def _write_new_file(self, data: pd.DataFrame, output_path: str) -> None:
        """Write data to a new Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Data"
        
        # Write headers
        for col_idx, column in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        
        # Write data
        for idx, row in data.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=idx + 2, column=col_idx, value=value)
        
        # Apply formatting
        self._apply_formatting(ws)
        
        # Save
        wb.save(output_path)
    
    def _apply_formatting(self, worksheet) -> None:
        """Apply red and blue theme formatting"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data row alternating colors
        data_fill_1 = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # Light blue
        data_fill_2 = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # Light red
        
        for row_idx in range(2, worksheet.max_row + 1):
            fill = data_fill_1 if row_idx % 2 == 0 else data_fill_2
            for cell in worksheet[row_idx]:
                cell.fill = fill


class ExcelProcessorService:
    """Main service class that orchestrates the processing workflow"""
    
    def __init__(self):
        self.reader = ExcelReader()
        self.processor = DataProcessor()
        self.writer = ExcelWriter()
    
    def process_file(self, input_path: str, output_path: str, template_path: Optional[str] = None, progress_callback: Optional[Callable] = None) -> Dict[str, Any]:
        """Main method to process Excel file"""
        try:
            if progress_callback:
                progress_callback(10, "Starting file processing...")
            
            # Prefer matrix-style parser if day headers are like '1 Mon'
            try:
                processed_data = self.processor.process_matrix_attendance(input_path, progress_callback)
            except Exception as matrix_err:
                logger.info(f"Matrix parser fallback: {matrix_err}")
                # Fallback to attendance parser (legacy format)
                if 'Report (1).xls' in input_path or self._is_attendance_file(input_path):
                    processed_data = self.processor.process_attendance_file(input_path, progress_callback)
                else:
                    # Final fallback to general wide-format processor
                    data = self.reader.read_excel(input_path, progress_callback)
                    processed_data = self.processor.process_data(data, progress_callback)
            
            # Write to output file
            self.writer.write_excel(processed_data, output_path, template_path, progress_callback)
            
            if progress_callback:
                progress_callback(100, f"Processing completed! {len(processed_data)} records processed.")
            
            return {
                'success': True,
                'input_rows': len(processed_data),
                'output_rows': len(processed_data),
                'output_path': output_path
            }
            
        except Exception as e:
            logger.error(f"Processing failed: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def _is_attendance_file(self, file_path: str) -> bool:
        """Check if the file is an attendance file based on content"""
        try:
            import pandas as pd
            
            # Try to read the file with pandas
            try:
                df = pd.read_excel(file_path, engine='openpyxl', header=None)
            except Exception:
                df = pd.read_excel(file_path, header=None)
            
            # Check if cell A9 (row 8, column 0) contains period information or similar
            if len(df) > 8:
                period_cell = df.iloc[8, 0]
                period_text = str(period_cell) if period_cell is not None else ''
                if any(keyword in period_text for keyword in ['Period:', 'PERIOD', 'period']):
                    return True
            
            # Check if row 11 (index 10) has date-like headers starting from column F (index 5)
            if len(df) > 10 and len(df.columns) > 5:
                header_row = df.iloc[10]
                sample_cells = [header_row[i] for i in range(5, min(len(header_row), 15))]
                for cell in sample_cells:
                    text = str(cell)
                    if any(char.isdigit() for char in text) and ('/' in text or '-' in text):
                        return True
                    
            return False
        except Exception:
            return False 